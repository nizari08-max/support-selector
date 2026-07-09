"""Phase 1 Excel output writer — clean, self-contained 2-sheet workbook.

Produces a fresh workbook (it does NOT mutate or copy the source workbook, so it
never drags along coversheets, notes, or merged-cell headers from the original
file):

  Sheet 1 "Classified Line List" — the original line-list rows (original headers
    and original cell values) plus three appended Phase 1 columns:
        CRITICALITY LEVEL      (I / II / III)
        CLASSIFICATION REASON
        DATA QUALITY FLAG

  Sheet 2 "Classification Report" — totals, criticality counts, data-detection
    assumptions, columns not detected, warnings, and the classification criteria
    summary.

No CN, P&ID, candidate-matching, or Phase 2 content is written.
"""

from __future__ import annotations

from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PHASE1_HEADERS = ["CRITICALITY LEVEL", "CLASSIFICATION REASON", "DATA QUALITY FLAG"]

LEVEL_TO_ROMAN = {"Level 1": "I", "Level 2": "II", "Level 3": "III"}

COLORS = {
    "LEVEL_I": "FFFF9999",    # red
    "LEVEL_II": "FFFFC266",   # orange
    "LEVEL_III": "FF92D050",  # green
    "EXCLUDED": "FFD9D9D9",   # grey
    "REVIEW": "FFFFFF99",     # yellow
    "HEADER": "FF1F4E79",     # navy
    "SECTION": "FF2E75B6",    # medium blue
    "TITLE": "FF1F4E79",
}

FONT_WHITE = Font(color="FFFFFFFF", bold=True)
FONT_BOLD = Font(bold=True)
THIN = Side(style="thin", color="FF808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_phase1_output(output_path: str, sheet1_df: pd.DataFrame, report: dict) -> None:
    """Write the clean 2-sheet Phase 1 workbook.

    sheet1_df: original line-list columns with the three PHASE1_HEADERS columns
               appended as the final three columns (CRITICALITY LEVEL already
               rendered as I/II/III).
    report:    dict of summary values for the Classification Report sheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classified Line List"
    _write_line_list_sheet(ws, sheet1_df)

    report_ws = wb.create_sheet("Classification Report")
    _write_report_sheet(report_ws, report)

    try:
        wb.save(output_path)
    except PermissionError as exc:  # pragma: no cover - environment specific
        raise PermissionError(
            f"Cannot write to '{output_path}'. Close it in Excel and retry."
        ) from exc


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Classified Line List
# ─────────────────────────────────────────────────────────────────────────────

def _write_line_list_sheet(ws, df: pd.DataFrame) -> None:
    columns = list(df.columns)
    n_cols = len(columns)

    # Header row
    for c_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=str(name))
        cell.fill = _fill(COLORS["HEADER"])
        cell.font = FONT_WHITE
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    try:
        level_col = columns.index("CRITICALITY LEVEL")
        dq_col = columns.index("DATA QUALITY FLAG")
    except ValueError:
        level_col = dq_col = None

    # Data rows
    for r_offset, (_, row) in enumerate(df.iterrows()):
        excel_row = r_offset + 2
        level_val = str(row.iloc[level_col]) if level_col is not None else ""
        dq_val = str(row.iloc[dq_col]) if dq_col is not None else ""
        row_fill = _row_fill(level_val, dq_val)

        for c_idx, name in enumerate(columns, start=1):
            value = row.iloc[c_idx - 1]
            value = "" if value is None or (isinstance(value, float) and pd.isna(value)) else value
            cell = ws.cell(row=excel_row, column=c_idx, value=value)
            cell.border = BORDER
            is_reason = name == "CLASSIFICATION REASON"
            cell.alignment = Alignment(
                horizontal="left" if is_reason else "center",
                vertical="top",
                wrap_text=is_reason,
            )
            if row_fill is not None:
                cell.fill = row_fill

    # Column widths
    for c_idx, name in enumerate(columns, start=1):
        letter = get_column_letter(c_idx)
        if name == "CLASSIFICATION REASON":
            ws.column_dimensions[letter].width = 72
        elif name == "DATA QUALITY FLAG":
            ws.column_dimensions[letter].width = 30
        elif name == "CRITICALITY LEVEL":
            ws.column_dimensions[letter].width = 16
        else:
            ws.column_dimensions[letter].width = 16

    ws.freeze_panes = "A2"
    if len(df) >= 0 and n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{len(df) + 1}"


def _row_fill(level_roman: str, data_quality: str) -> Optional[PatternFill]:
    dq = str(data_quality or "").strip()
    if dq.upper().startswith("SCOPE:"):
        return _fill(COLORS["EXCLUDED"])
    if str(level_roman).strip() == "I":
        return _fill(COLORS["LEVEL_I"])
    if str(level_roman).strip() == "II":
        return _fill(COLORS["LEVEL_II"])
    if str(level_roman).strip() == "III":
        return _fill(COLORS["LEVEL_III"])
    if dq:
        return _fill(COLORS["REVIEW"])
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Classification Report
# ─────────────────────────────────────────────────────────────────────────────

def _write_report_sheet(ws, report: dict) -> None:
    not_detected = report.get("columns_not_detected") or []
    not_detected_txt = ", ".join(not_detected) if not_detected else "none"
    warnings = report.get("warnings") or []

    rows: list[tuple] = [
        ("Critical Line List Classifier — Classification Report", None, "title"),
        (None, None, None),
        ("Metric", "Value", "header"),
        ("Total lines in file", report.get("total_lines", 0), None),
        ("Lines in scope", report.get("in_scope_lines", 0), None),
        ("Lines excluded (vendor / client / licensor)", report.get("excluded_lines", 0), None),
        (None, None, None),
        ("Criticality I — rigorous stress analysis", report.get("level_1_count", 0), "l1"),
        ("Criticality II — normal stress analysis", report.get("level_2_count", 0), "l2"),
        ("Criticality III — visual / span check", report.get("level_3_count", 0), "l3"),
        ("Missing data (cannot classify)", report.get("missing_data_count", 0), None),
        ("Ambiguous data (unknown material code)", report.get("ambiguous_count", 0), None),
        (None, None, None),
        ("Data Detection & Key Assumptions", None, "section"),
        ("Source worksheet", report.get("sheet_name", ""), None),
        ("Header row (Excel)", report.get("header_excel_row", ""), None),
        ("Units row skipped (Excel)", report.get("units_row_skipped", "none"), None),
        ("Size unit", report.get("size_unit_label", ""), None),
        ("Scope handling", report.get("scope_label", ""), None),
        ("Equipment detection mode", report.get("equipment_mode", ""), None),
        ("Columns mapped", report.get("columns_mapped", 0), None),
        ("Columns not detected", not_detected_txt, None),
    ]

    rows.append((None, None, None))
    rows.append(("Warnings & Data-Quality Notes", None, "section"))
    if warnings:
        for w in warnings:
            rows.append((str(w), None, "wrap"))
    else:
        rows.append(("No warnings — all columns detected and rows classified.", None, None))

    rows.append((None, None, None))
    rows.append(("Classification Criteria (JESA Stress Critical Line List)", None, "section"))
    for line in _CRITERIA_LINES:
        rows.append((line, None, "wrap"))

    rows.append((None, None, None))
    rows.append(("Color Legend", None, "section"))
    for label, color_key in (
        ("Criticality I", "LEVEL_I"),
        ("Criticality II", "LEVEL_II"),
        ("Criticality III", "LEVEL_III"),
        ("Excluded (out of scope)", "EXCLUDED"),
        ("Needs review (missing / ambiguous)", "REVIEW"),
    ):
        rows.append((label, None, ("legend", color_key)))

    for r_idx, entry in enumerate(rows, start=1):
        label, value, style = entry
        c1 = ws.cell(row=r_idx, column=1, value=label)
        c2 = ws.cell(row=r_idx, column=2, value=value)

        if style == "title":
            c1.font = Font(bold=True, size=14, color=COLORS["TITLE"])
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=2)
        elif style == "header":
            for cc in (c1, c2):
                cc.fill = _fill(COLORS["SECTION"])
                cc.font = FONT_WHITE
                cc.alignment = Alignment(horizontal="center")
                cc.border = BORDER
        elif style == "section":
            c1.font = FONT_BOLD
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=2)
        elif style == "wrap":
            c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=2)
        elif isinstance(style, tuple) and style[0] == "legend":
            fill = _fill(COLORS[style[1]])
            c1.fill = fill
            c2.fill = fill
            c1.border = BORDER
            c2.border = BORDER
        elif label is not None:
            c1.border = BORDER
            c2.border = BORDER
            if value is not None:
                c1.font = FONT_BOLD

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 20


_CRITERIA_LINES = [
    "Deterministic 5-step waterfall (first match wins; no later step overrides an earlier one):",
    "  1. Non-metallic material (FRP / GRE / HDPE) → Level I unconditionally.",
    "  2. Exception flags (relief line, jacketed, vibration, settlement, expansion joint,",
    "     vacuum, PSV high-pressure, etc.) → Level I.",
    "  3. Connected equipment: centrifugal pump / strain-sensitive → Chart 3; other",
    "     equipment nozzle → Chart 4 (small-bore < 2\" capped at Level II).",
    "  4. Material chart lookup by size band and design temperature — Chart 1 (carbon /",
    "     low-alloy steel) or Chart 2 (stainless / duplex / alloy).",
    "  5. Missing size, design temperature, or material → left unclassified and flagged",
    "     for engineer review.",
    "Upgrade rules may raise criticality for corrosive service, cold / cryogenic service,",
    "high-pressure steam condensate, and 904L alloys on aggressive service.",
    "Levels: I = rigorous flexibility/stress analysis, II = normal analysis, III = visual",
    "or span check only.",
]


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)
